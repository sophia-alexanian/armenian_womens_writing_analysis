from openpyxl import load_workbook, Workbook

def copy_column_to_new_excel(source_file, column_letter, destination_file):
    # Load the source workbook and select the first sheet
    source_wb = load_workbook(source_file)
    source_sheet = source_wb.active

    # Create a new workbook
    dest_wb = Workbook()
    dest_sheet = dest_wb.active

    # Get the column index from the letter
    col_idx = source_sheet[column_letter + "1"].column

    # Copy data from the column
    for row in source_sheet.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
        dest_sheet.append([row[0]])

    # Save the new workbook
    dest_wb.save(destination_file)
    print(f"Column {column_letter} copied to {destination_file} successfully.")

# Example usage
source_file = "Normalized_Armenian_Womens_Articles.xlsx"  # Path to your source Excel file
destination_file = "keywords.xlsx"  # Path to save the new Excel file
column_letter = "D"  # Column to copy (e.g., 'A')

copy_column_to_new_excel(source_file, column_letter, destination_file)
