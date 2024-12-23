import openai
import openpyxl
import time
import requests
import json

# Set your OpenAI API key
openai.api_key = "INSERT-YOUR-KEY-HERE"  #keys are private
url = "https://api.openai.com/v1/chat/completions"
headers = {
    "Authorization": f"Bearer {openai.api_key}",
    "Content-Type": "application/json"
}

# Define the function to determine gender using ChatGPT
def determine_gender(name):
    try:
        messages = [
            {"role": "system", "content": "You are an assistant that determines the likely gender of a name."},
            {"role": "user", "content": f"What is the gender of the name 'Ara'? Reply with 'male', 'female', or 'uncertain'."},
            {"role": "assistant", "content": "male"},
            {"role": "user", "content": f"What is the gender of the name 'Arpi'? Reply with 'male', 'female', or 'uncertain'."},
            {"role": "assistant", "content": "female"},
            {"role": "user", "content": f"What is the gender of the name 'Ani'? Reply with 'male', 'female', or 'uncertain'."},
            {"role": "assistant", "content": "female"},
            {"role": "user", "content": f"What is the gender of the name 'Sarine'? Reply with 'male', 'female', or 'uncertain'."},
            {"role": "assistant", "content": "female"},
            {"role": "user", "content": f"What is the gender of the name 'Murad'? Reply with 'male', 'female', or 'uncertain'."},
            {"role": "assistant", "content": "male"},
            {"role": "user", "content": f"What is the gender of the name '{name}'? Reply with 'male', 'female', or 'uncertain'."},
        ]
        response = openai.ChatCompletion.create(
            model= "chatgpt-4o-latest",  # Use a valid chat model
            messages=messages,
            max_tokens=10
        )
        gender = response.choices[0].message['content'].strip()
        return gender
    except Exception as e:
        print(f"Error determining gender for {name}: {e}")
        return "Error"

def determine_gender_with_retry(name, retries=3, delay=5):
    for attempt in range(retries):
        try:
            data = {
                "model": "gpt-3.5-turbo",  # or "gpt-4"
                "messages": [
                    {"role": "system", "content": "You are an assistant that determines the likely gender of a name."},
                    {"role": "user", "content": f"What is the gender of the name '{name}'? Reply with 'male', 'female', or 'uncertain'."}
                ]
            }

            response = requests.post(url, headers=headers, data=json.dumps(data))
            response_data = response.json()  # Parse the JSON response

            if response.status_code == 200:
                gender = response_data['choices'][0]['message']['content'].strip()
                return gender
            else:
                print(f"Error: {response_data.get('error', {}).get('message', 'Unknown error')}")
                return "Error"
        except Exception as e:
            if attempt < retries - 1:
                print(f"Error determining gender for {name}: {e}. Retrying in {delay} seconds...")
                time.sleep(delay)
            else:
                print(f"Failed to determine gender for {name} after {retries} attempts.")
                return "Error"

# Load the Excel workbook and sheet
input_file = "unique_author_names.xlsx"  # Replace with your Excel file name
output_file = "unique_gendered_author_names.xlsx"
sheet_name = "Sheet1"  # Replace with the name of your sheet

wb = openpyxl.load_workbook(input_file)
sheet = wb[sheet_name]

# Iterate through rows and determine gender
name_column = 1  # Replace with the column index of names (1-based)
gender_column = 2  # Replace with the desired column index for gender (1-based)

for row in range(2, sheet.max_row + 1):
    name = sheet.cell(row=row, column=name_column).value
    if name:
        gender = determine_gender(name)
        sheet.cell(row=row, column=gender_column, value=gender)

# Save the updated workbook
wb.save(output_file)
print(f"Updated sheet saved to {output_file}")


