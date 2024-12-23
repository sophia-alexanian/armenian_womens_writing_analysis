import newspaper
import openpyxl
from newspaper import Article
from bs4 import BeautifulSoup
import requests
import random
import time

# Initialize a new Excel workbook
wb = openpyxl.Workbook()
sheet = wb.active

# Set column headers
sheet.cell(row=1, column=1).value = "Title"
sheet.cell(row=1, column=2).value = "Author"
sheet.cell(row=1, column=3).value = "Publish Date"
sheet.cell(row=1, column=4).value = "Keywords"
sheet.cell(row=1, column=5).value = "URL"

# Base URLs for categories and their pagination limits
categories_with_pages = {
    "https://armenianweekly.com/diaspora/community-news/": 66,
    "https://armenianweekly.com/news/reports/" : 2,
    "https://armenianweekly.com/news/briefs/" : 14, 
     "https://armenianweekly.com/news/analysis/" : 1, 
    "https://armenianweekly.com/news/roundups/" : 3,
    "https://armenianweekly.com/opinion/a-word-from-the-editor/" : 2,
    "https://armenianweekly.com/opinion/commentary/" : 7,
    "https://armenianweekly.com/opinion/op-eds/" : 53,
    "https://armenianweekly.com/opinion/letters/" : 8,
    "https://armenianweekly.com/opinion/open-letters/" : 3,
    "https://armenianweekly.com/columns/we-the-armenians/" : 2,
    "https://armenianweekly.com/columns/empty-barrel/" : 1,
    "https://armenianweekly.com/columns/in-sight/" : 29,
    "https://armenianweekly.com/author/harut-sassounian/" : 77,
    "https://armenianweekly.com/columns/hye-key/" : 5,
    "https://armenianweekly.com/columns/in-pursuit-of-home/" : 1,
    "https://armenianweekly.com/columns/armenian-sports-block/" : 3,
    "https://armenianweekly.com/columns/tashjians-take/" : 9,
    "https://armenianweekly.com/columns/into-the-archives/" : 1,
    "https://armenianweekly.com/author/ck-garabed/" : 25,
    "https://armenianweekly.com/columns/victorias-voice/" : 2,
    "https://armenianweekly.com/author/garen-yegparian/" : 48,
    "https://armenianweekly.com/columns/keeping-the-faith/" : 8,
    "https://armenianweekly.com/diaspora/statements/" : 21,
    "https://armenianweekly.com/youth/youth-opinion/" : 2,
    "https://armenianweekly.com/diaspora/anca-news/" : 74,
    "https://armenianweekly.com/culture/film/" : 8,
    "https://armenianweekly.com/culture/photography/" : 3,
    "https://armenianweekly.com/culture/sports/" : 5,
    "https://armenianweekly.com/culture/the-making-of/" : 2,
    "https://armenianweekly.com/culture/music/" : 12,
    "https://armenianweekly.com/culture/travel/" : 6,
    "https://armenianweekly.com/literary-corner/fiction/" : 2,
    "https://armenianweekly.com/literary-corner/memoir/" : 2,
    "https://armenianweekly.com/literary-corner/poetry/" : 9,
    "https://armenianweekly.com/literary-corner/reviews/" : 12
}

# Initialize row counter for Excel
row = 2

# Initialize a set to track processed URLs
processed_urls = set()

# Set headers to simulate a browser request
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36"
}

# Iterate through each category and its pages
for base_url, max_pages in categories_with_pages.items():
    for page_num in range(1, max_pages + 1):
        # Construct the paginated URL
        paginated_url = f"{base_url}page/{page_num}/" if page_num > 1 else base_url
        print(f"Scraping page: {paginated_url}")

        # Random delay between 1 and 3 seconds before making a request
        time.sleep(random.uniform(1, 3))

        # Fetch the page HTML with the custom headers
        try:
            response = requests.get(paginated_url, headers=headers)
            response.raise_for_status()
        except Exception as e:
            print(f"Failed to fetch {paginated_url}: {e}")
            continue

        # Parse the page content using BeautifulSoup
        soup = BeautifulSoup(response.content, 'html.parser')

        # Find all article links using a broader CSS selector
        article_links = soup.select('a[rel="bookmark"]')  # Adjusted to select all <a> tags with rel="bookmark"
        if not article_links:
            print(f"No articles found on {paginated_url}")
            continue

        # Loop through each article link
        for link in article_links:
            article_url = link['href']

            # Skip comment URLs
            if article_url.endswith('#comments'):
                continue

            # Check for duplicate URLs
            if article_url in processed_urls:
                print(f"Skipping duplicate article: {article_url}")
                continue

            print(f"Processing article: {article_url}")

            try:
                # Add the URL to the processed set
                processed_urls.add(article_url)

                # Download and parse the article using newspaper3k
                article = Article(article_url)
                article.download()
                article.parse()
                article.nlp()

                # Write article details to Excel
                sheet.cell(row=row, column=1).value = article.title
                sheet.cell(row=row, column=2).value = article.authors[0] if article.authors else "N/A"
                sheet.cell(row=row, column=3).value = (
                    article.publish_date.strftime('%Y-%m-%d') if article.publish_date else "N/A"
                )
                sheet.cell(row=row, column=4).value = ', '.join(article.keywords)
                sheet.cell(row=row, column=5).value = article_url

                # Increment the row counter
                row += 1
            except Exception as e:
                print(f"Failed to process article {article_url}: {e}")
                continue

# Save the workbook
output_path = r"..\ArmenianWomensWritingAnalysis\scrapers\ArmenianWeeklyScraper\Armenian_Weekly_Articles.xlsx"
wb.save(output_path)

print(f"Scraping completed. Data saved to {output_path}")
