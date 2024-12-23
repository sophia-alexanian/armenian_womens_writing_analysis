import openpyxl
from newspaper import Article
from bs4 import BeautifulSoup
import requests
import random
import time

# Initialize a new Excel workbook
wb = openpyxl.Workbook()
sheet = wb.active

# Set column headers
sheet.cell(row=1, column=1).value = "Title"
sheet.cell(row=1, column=2).value = "Author"
sheet.cell(row=1, column=3).value = "Publish Date"
sheet.cell(row=1, column=4).value = "Keywords"
sheet.cell(row=1, column=5).value = "URL"

categories_with_pages = {
    "https://evnreport.com/series-category/evn-security-report/": 4,
    "https://evnreport.com/series-category/politics/" : 37,
    "https://evnreport.com/series-category/opinion/" : 21,
    "https://evnreport.com/series-category/column/" : 4,
    "https://evnreport.com/series-category/spotlight-karabakh/" : 19,
    "https://evnreport.com/series-category/raw-unfiltered/" : 35,
    "https://evnreport.com/series-category/arts-and-culture/" : 11,
    "https://evnreport.com/series-category/et-cetera/" : 8,
    "https://evnreport.com/series-category/reviews/" : 4,
    "https://evnreport.com/series-category/creative-tech/" : 7,
    "https://evnreport.com/series-category/law-society/" : 6,
    "https://evnreport.com/series-category/economy/" : 9,
    "https://evnreport.com/series-category/readers-forum/" : 4
}

# Initialize row counter for Excel
row = 2

# Initialize a set to track processed URLs
processed_urls = set()

# Set headers to simulate a browser request
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36"
}

# Maximum number of pages to scrape
max_pages = 10  # Adjust based on the actual number of pages

# Iterate through each category
for base_url, max_pages in categories_with_pages.items():
    # Iterate through each page
    for page_num in range(1, max_pages + 1):
        # Construct the paginated URL
        paginated_url = base_url if page_num == 1 else f"{base_url}page/{page_num}/"
        print(f"Scraping page: {paginated_url}")

        # Random delay between 1 and 3 seconds before making a request
        time.sleep(random.uniform(1, 3))

        # Fetch the page HTML with the custom headers
        try:
            response = requests.get(paginated_url, headers=headers)
            response.raise_for_status()
        except Exception as e:
            print(f"Failed to fetch {paginated_url}: {e}")
            continue

        # Parse the page content using BeautifulSoup
        soup = BeautifulSoup(response.content, 'html.parser')

        # Find all article links using the specified HTML structure
        article_links = soup.select('h3.jeg_post_title > a')
        if not article_links:
            print(f"No articles found on {paginated_url}")
            continue

        # Loop through each article link
        for link in article_links:
            article_url = link['href']

            # Check for duplicate URLs
            if article_url in processed_urls:
                print(f"Skipping duplicate article: {article_url}")
                continue

            print(f"Processing article: {article_url}")

            try:
                # Add the URL to the processed set
                processed_urls.add(article_url)

                # Download and parse the article using newspaper3k
                article = Article(article_url)
                article.download()
                article.parse()
                article.nlp()

                # Write article details to Excel
                sheet.cell(row=row, column=1).value = article.title
                sheet.cell(row=row, column=2).value = article.authors[0] if article.authors else "N/A"
                sheet.cell(row=row, column=3).value = (
                    article.publish_date.strftime('%Y-%m-%d') if article.publish_date else "N/A"
                )
                sheet.cell(row=row, column=4).value = ', '.join(article.keywords)
                sheet.cell(row=row, column=5).value = article_url

                # Increment the row counter
                row += 1
            except Exception as e:
                print(f"Failed to process article {article_url}: {e}")
                continue

# Save the workbook
output_path = r"..\ArmenianWomensWritingAnalysis\scrapers\EVNReportScraper\EVN_Report_Articles.xlsx"
wb.save(output_path)

print(f"Scraping completed. Data saved to {output_path}")