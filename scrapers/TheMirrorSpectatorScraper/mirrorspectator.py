import newspaper
import openpyxl
from newspaper import Article
from bs4 import BeautifulSoup
import requests
import random
import time

# Initialize a new Excel workbook
wb = openpyxl.Workbook()
sheet = wb.active

# Set column headers
sheet.cell(row=1, column=1).value = "Title"
sheet.cell(row=1, column=2).value = "Author"
sheet.cell(row=1, column=3).value = "Publish Date"
sheet.cell(row=1, column=4).value = "Keywords"
sheet.cell(row=1, column=5).value = "URL"

# Base URLs for categories and their pagination limits
categories_with_pages = {
    "https://mirrorspectator.com/archives/": 1550
}

# Initialize row counter for Excel
row = 2

# Initialize a set to track processed URLs
processed_urls = set()

# Set headers to simulate a browser request
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36"
}

# Author extraction logic
def extract_author(article_soup):
    # Check for the second type: div.td-post-author-name
    author_div = article_soup.select_one('.td-post-author-name a')
    if author_div:
        author_name = author_div.get_text(strip=True)
        if author_name == "The Armenian Mirror-Spectator":
            # Check for the first type: <p> with "By ..."
            author_paragraph = article_soup.find('p', text=lambda x: x and x.strip().startswith("By "))
            if author_paragraph:
                # Extract author name from the paragraph
                return author_paragraph.get_text(strip=True).replace("By ", "").strip()
        return author_name

    # Fallback to checking for the first type directly if the second type is not found
    author_paragraph = article_soup.find('p', text=lambda x: x and x.strip().startswith("By "))
    if author_paragraph:
        return author_paragraph.get_text(strip=True).replace("By ", "").strip()

    # Default to "N/A" if no author information is found
    return "N/A"

# Iterate through each category and its pages
for base_url, max_pages in categories_with_pages.items():
    for page_num in range(1, max_pages + 1):
        # Construct the paginated URL
        paginated_url = f"{base_url}page/{page_num}/" if page_num > 1 else base_url
        print(f"Scraping page: {paginated_url}")

        # Random delay between 1 and 3 seconds before making a request
        time.sleep(random.uniform(1, 3))

        # Fetch the page HTML with the custom headers
        try:
            response = requests.get(paginated_url, headers=headers)
            response.raise_for_status()
        except Exception as e:
            print(f"Failed to fetch {paginated_url}: {e}")
            continue

        # Parse the page content using BeautifulSoup
        soup = BeautifulSoup(response.content, 'html.parser')

        # Find all article links using a broader CSS selector
        article_links = soup.select('a[rel="bookmark"]')  # Adjusted to select all <a> tags with rel="bookmark"
        if not article_links:
            print(f"No articles found on {paginated_url}")
            continue

        # Loop through each article link
        for link in article_links:
            article_url = link['href']

            # Skip comment URLs
            if article_url.endswith('#comments'):
                continue

            # Check for duplicate URLs
            if article_url in processed_urls:
                print(f"Skipping duplicate article: {article_url}")
                continue

            print(f"Processing article: {article_url}")

            try:
                # Add the URL to the processed set
                processed_urls.add(article_url)

                # Download and parse the article using newspaper3k
                article = Article(article_url)
                article.download()
                article.parse()
                article.nlp()

                # Fetch the article's soup for additional parsing
                article_soup = BeautifulSoup(article.html, 'html.parser')

                # Extract author using the new logic
                author_name = extract_author(article_soup)

                # Write article details to Excel
                sheet.cell(row=row, column=1).value = article.title
                sheet.cell(row=row, column=2).value = author_name
                sheet.cell(row=row, column=3).value = (
                    article.publish_date.strftime('%Y-%m-%d') if article.publish_date else "N/A"
                )
                sheet.cell(row=row, column=4).value = ', '.join(article.keywords)
                sheet.cell(row=row, column=5).value = article_url

                # Increment the row counter
                row += 1
            except Exception as e:
                print(f"Failed to process article {article_url}: {e}")
                continue

# Save the workbook
output_path = r"..\ArmenianWomensWritingAnalysis\scrapers\TheMirrorSpectatorScraper\The_Mirror_Spectator_Articles.xlsx"
wb.save(output_path)

print(f"Scraping completed. Data saved to {output_path}")