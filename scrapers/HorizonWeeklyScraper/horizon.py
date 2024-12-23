import requests
from bs4 import BeautifulSoup
import openpyxl
from newspaper import Article
import time
import random
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

# Initialize a new Excel workbook
wb = openpyxl.Workbook()
sheet = wb.active

# Automatically manage ChromeDriver with Selenium Manager
driver = webdriver.Chrome()

# Set column headers
sheet.cell(row=1, column=1).value = "Title"
sheet.cell(row=1, column=2).value = "Author"
sheet.cell(row=1, column=3).value = "Publish Date"
sheet.cell(row=1, column=4).value = "Keywords"
sheet.cell(row=1, column=5).value = "URL"

# Set headers to simulate a browser request
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36"
}

# Base URL for the category with individual authors
base_url = "https://horizonweekly.ca/en/category/opinion/"

# Initialize row counter for Excel
row = 2
page_number = 1

# Initialize a set to track processed URLs
processed_urls = set()

# Function to simulate random delays between requests
def random_delay():
    time.sleep(random.uniform(2, 5))  # Sleep between 2 to 5 seconds

# Start scraping from the first page
current_url = base_url

while current_url:
    print(f"Scraping page: {current_url}")

    try:
        # Fetch the page content
        response = requests.get(current_url, headers=headers)
        response.raise_for_status()
    except Exception as e:
        print(f"Failed to fetch {current_url}: {e}")
        break  # Stop scraping if we can't fetch the page

    # Parse the page using BeautifulSoup
    soup = BeautifulSoup(response.content, 'html.parser')

    # Find all article links
    article_links = soup.select('h4.article-names > a')  # Correct the selector if needed
    if not article_links:
        print(f"No articles found on {current_url}")
        break

    # Loop through each article link
    for link in article_links:
        article_url = link['href']

        # Check for duplicate URLs
        if article_url in processed_urls:
            print(f"Skipping duplicate article: {article_url}")
            continue

        print(f"Processing article: {article_url}")

        try:
            # Add the URL to the processed set
            processed_urls.add(article_url)

            # Download and parse the article using newspaper3k
            article = Article(article_url)
            article.download()
            article.parse()
            article.nlp()

            # Write article details to Excel
            sheet.cell(row=row, column=1).value = article.title
            sheet.cell(row=row, column=3).value = (article.publish_date.strftime('%Y-%m-%d') if article.publish_date else "N/A")
            sheet.cell(row=row, column=4).value = ', '.join(article.keywords)
            sheet.cell(row=row, column=5).value = article_url

            # Special case for finding author using Selenium
            driver.get(article_url)  # Navigate to the article URL
            html_content = driver.page_source  # Get the HTML content of the page
            article_soup = BeautifulSoup(html_content, 'html.parser')
            author_pattern = re.compile(r'^\s*(BY|By)\s+([A-Z][a-z]+(?:\s[A-Z][a-z]+)*)', re.IGNORECASE)

            author = "N/A"
            for tag in article_soup.find_all(['strong', 'p']):
                text = tag.get_text(strip=True)
                match = author_pattern.match(text)
                if match:
                    author = match.group(2)  # Extract the author's name
                    break

            sheet.cell(row=row, column=2).value = author

            # Increment the row counter
            row += 1

            # Random delay between requests
            random_delay()

        except Exception as e:
            print(f"Failed to process article {article_url}: {e}")
            continue

    # Move to the next page
    page_number += 1
    current_url = f"{base_url}page/{page_number}/"

# Save the workbook
output_path = r"..\ArmenianWomensWritingAnalysis\scrapers\HorizonWeeklyScraper\Horizon_Weekly_Articles.xlsx"

try:
    wb.save(output_path)
    print(f"Scraping completed. Data saved to {output_path}")
finally:
    # Close the Selenium driver to release resources
    driver.quit()

